###############################################
from sklearn.linear_model import LinearRegression

from sklearn.ensemble import RandomForestRegressor
import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error  # 正确导入方式
from sklearn.metrics import r2_score  # 必须导入！
import matplotlib.pyplot as plt
import seaborn as sns

from scipy.stats import spearmanr
from scipy.stats import pearsonr

DATA = pd.read_excel('C:/Users/32684/.spyder-py3/20250418自定义损失/Atherosclerosis_data.xlsx', header=None)

DATA1 = DATA.values
print(DATA1)

Y=DATA1[1:,1]
print(Y)

#X=DATA1[1:,[3,4]]
X1=DATA1[1:,2:-1]
print(X1)

print(DATA1[0,2:-1])

X=X1[:,[0,1,2,4,7]]

Sex=DATA1[1:,0]
Target=DATA1[1:,-1]

X_male= X[np.where(Sex == 0)[0],:]
Y_male= Y[np.where(Sex == 0)[0]]
Target_male=Target[np.where(Sex == 0)[0]]

X_male_h = X_male[np.where(Target_male == 0)[0],:]
Y_male_h = Y_male[np.where(Target_male == 0)[0]]
X_male_noh = X_male[np.where(Target_male == 1)[0],:]
Y_male_noh = Y_male[np.where(Target_male == 1)[0]]

#★★★★★★★★★★★★★★★★★★★★★★★现在测试男性的数据
X_model=X_male_h
Y_model=Y_male_h
#print(np.size(Y_model))


X_train=X_model[::2]  # 从索引0开始，步长为2
y_train=Y_model[::2]

X_test=X_model[1::2]  # 从索引1开始，步长为2 
y_test=Y_model[1::2]

import xgboost as xgb
from xgboost import XGBRegressor

##   可视化测试训练和测试，利用固定参数
# 创建XGBoost回归模型实例
xgb_reg = xgb.XGBRegressor(
    objective='reg:squarederror',  # 回归任务的目标函数
    n_estimators=80,              # 树的数量
    learning_rate=0.1,             # 学习率
    max_depth=2,                   # 树的最大深度
)
#type(0.4)
xgb_reg.fit(X_train, y_train)
train_predictions = xgb_reg.predict(X_train)
test_predictions = xgb_reg.predict(X_test)

#model2是基于X_train和y_train，构造mpa模型
from sklearn.linear_model import LinearRegression
model2 = LinearRegression()
model2.fit(y_train.reshape(-1, 1),train_predictions)
#构造训练集的mpa
train_predictions_mpa = model2.predict(y_train.reshape(-1, 1))
#构造测试集的mpa
test_predictions_mpa = model2.predict(y_test.reshape(-1, 1))

z_train=train_predictions-train_predictions_mpa
z_test=test_predictions-test_predictions_mpa
'''
xgb_reg.fit(X_train, y_train+z_train)
train_predictions = xgb_reg.predict(X_train)
plt.scatter(y_train, train_predictions)
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='--', label='y=x')
'''

mean_absolute_error(y_test, test_predictions)################################MAE--5.161
mean_squared_error(y_test, test_predictions)#train_mse#######################MSE--40.88
corr2, p_value2 = spearmanr(y_test, test_predictions)########################相关0.834
corr2
r2_score(y_test, test_predictions)###########################################R2=0.568
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
test_result = test_predictions-y_test
    # 找到小于50的样本的索引
indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)#############小于50--0.51
    # 找到大于等于50的样本的索引
indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)##############大于50---0.52


plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
plt.subplot(2, 3, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions,s=10,color='b', alpha=0.2)
corr_coef = np.corrcoef(np.array(y_train, dtype=float), train_predictions.ravel() )[0, 1]
#mean_squared_error(y_train, train_predictions)
modela = LinearRegression()
modela.fit(y_train.reshape(-1, 1), train_predictions)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}\nr = {corr_coef:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Arterial biological age')
plt.title(f'Female (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions-y_train,s=10,color='b', alpha=0.2)
modela.fit(y_train.reshape(-1, 1), train_predictions-y_train)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
#plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Residual')
plt.xlim(20, 90)
plt.ylim(-35, 35)
plt.title(f'Female (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 3)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(train_predictions-y_train, kde=True, bins=20, color='b', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Pre-Residual')
plt.xlim(-35, 35)
plt.ylim(0, 0.075)
plt.title(f'Female (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions,s=10,color='c', alpha=0.2)
corr_coef = np.corrcoef(np.array(y_test, dtype=float), test_predictions.ravel() )[0, 1]
modela.fit(y_test.reshape(-1, 1), test_predictions)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}\nr = {corr_coef:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Arterial biological age')
plt.title(f'Female (XGBoost, test, n={len(y_test)})')

plt.subplot(2, 3, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions-y_test,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions-y_test)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
#plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlim(20, 90)
plt.xlabel('Chronological age')
plt.ylabel('Pre-Residual')
plt.xlim(20, 90)
plt.ylim(-35, 35)
plt.title(f'Female (XGBoost, test, n={len(y_test)})')

plt.subplot(2, 3, 6)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(test_predictions-y_test, kde=True, bins=20, color='c', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Pre-Residual')
plt.xlim(-35, 35)
plt.ylim(0, 0.075)
plt.title(f'Female (XGBoost, test, n={len(y_test)})')
plt.tight_layout()
plt.show()


#####--★★★★★★★★★★★★★★★★★★★★★★★★★★-L1结果
xgb_reg.fit(X_train, y_train+z_train)
train_predictions_l1 = xgb_reg.predict(X_train)
test_predictions_l1 = xgb_reg.predict(X_test)
'''
plt.scatter(y_train, train_predictions)
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='--', label='y=x')
'''
mean_absolute_error(y_test, test_predictions_l1)################################MAE--5.161
mean_squared_error(y_test, test_predictions_l1)#train_mse#######################MSE--40.88
corr2, p_value2 = spearmanr(y_test, test_predictions_l1)########################相关0.834
corr2
r2_score(y_test, test_predictions_l1)###########################################R2=0.568
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
test_result = test_predictions_l1-y_test
    # 找到小于50的样本的索引
indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)#############小于50--0.51
    # 找到大于等于50的样本的索引
indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)##############大于50---0.52


plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
plt.subplot(2, 3, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions_l1,s=10,color='b', alpha=0.2)
#mean_squared_error(y_train, train_predictions)
modela = LinearRegression()
modela.fit(y_train.reshape(-1, 1), train_predictions_l1)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Re-Arterial age')
plt.title(f'Male (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions_l1-y_train,s=10,color='b', alpha=0.2)
modela.fit(y_train.reshape(-1, 1), train_predictions_l1-y_train)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlabel('Chronological age')
plt.ylabel('Re-Residual')
plt.xlim(20, 90)
plt.ylim(-30, 30)
plt.title(f'Male (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 3)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(train_predictions_l1-y_train, kde=True, bins=20, color='b', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Re-Residual')
plt.xlim(-30, 30)
plt.ylim(0, 0.075)
plt.title(f'Male (XGBoost, train, n={len(y_train)})')

plt.subplot(2, 3, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions_l1,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions_l1)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Re-Arterial age')
plt.title(f'Male (XGBoost, test, n={len(y_test)})')

plt.subplot(2, 3, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions_l1-y_test,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions_l1-y_test)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlim(20, 90)
plt.xlabel('Chronological age')
plt.ylabel('Re-Residual')
plt.xlim(20, 90)
plt.ylim(-30, 30)
plt.title(f'Male (XGBoost, test, n={len(y_test)})')

plt.subplot(2, 3, 6)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(test_predictions_l1-y_test, kde=True, bins=20, color='c', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Re-Residual')
plt.xlim(-30, 30)
plt.ylim(0, 0.075)
plt.title(f'Male (XGBoost, test, n={len(y_test)})')
plt.tight_layout()
plt.show()



# **********寻找最优参数---------定义参数范围
'''
n_estimators = np.arange(1, 201, 10)      # [1, 11, 21, ..., 201]
learning_rates = np.arange(0.1, 1.0, 0.1) # [0.1, 0.2, ..., 0.9]
max_depths = np.arange(1, 10)             # [1, 2, ..., 9]
# 生成所有参数组合（3D网格）
grid = np.array(np.meshgrid(n_estimators, learning_rates, max_depths)).T.reshape(-1, 3)

len(grid)

train_test=np.zeros((len(grid), 2))  # 创建n行2列的数组
for i in range(len(grid)):  # range(20) 生成 0~19
    print(i)
    #i=1
    xgb_reg = xgb.XGBRegressor(
    objective='reg:squarederror',  # 回归任务的目标函数
    n_estimators=int(grid[i,0]),              # 树的数量
    learning_rate=grid[i,1],             # 学习率
    max_depth=int(grid[i, 2]))
    
    xgb_reg.fit(X_train, y_train)
    train_predictions = xgb_reg.predict(X_train)
    test_predictions = xgb_reg.predict(X_test)
    mean_squared_error(y_train, train_predictions)
    mean_squared_error(y_test, test_predictions)
    train_test[i,0]=mean_squared_error(y_train, train_predictions)
    train_test[i,1]=mean_squared_error(y_test, test_predictions)

from mpl_toolkits.mplot3d import Axes3D
fig = plt.figure(figsize=(10, 8))
ax = fig.add_subplot(111, projection='3d')
# 绘制散点图，c参数指定颜色，cmap指定颜色映射
scatter = ax.scatter(grid[:,0], grid[:,1], grid[:,2], c=train_test[:,1], cmap='viridis', s=50)
# 添加颜色条
cbar = plt.colorbar(scatter)
cbar.set_label('Value')

ax.set_xlabel('X')
ax.set_ylabel('Y')
ax.set_zlabel('Z')
plt.title('3D Scatter Plot with Color Mapping')
plt.show()

plt.scatter(train_test[:, 0], train_test[:, 0] / train_test[:, 1],s=10,color='b', alpha=0.2)
plt.scatter(train_test[:, 1], train_test[:, 0] / train_test[:, 1],s=10,color='r', alpha=0.2)
plt.grid(True)

plt.scatter(grid[:,2], train_test[:, 0] / train_test[:, 1],s=10,color='r', alpha=0.2)
plt.grid(True)
#plt.scatter(grid[:,0], train_test[:, 1],s=10,color='r', alpha=0.2)
plt.grid(True)


indices1 = np.where(train_test[:, 1] < 60)[0]
# 计算比值
ratio = train_test[:, 0] / train_test[:, 1]
# 查找比值 > 0.95 的索引
indices2 = np.where(ratio > 0.95)[0]
common_indices = np.intersect1d(indices1, indices2)


from mpl_toolkits.mplot3d import Axes3D
fig = plt.figure(figsize=(10, 8))
ax = fig.add_subplot(111, projection='3d')
scatter = ax.scatter(grid[common_indices,0], grid[common_indices,1], grid[common_indices,2], c=train_test[common_indices,1], cmap='viridis', s=50)
# 添加颜色条
cbar = plt.colorbar(scatter)
cbar.set_label('Value')
ax.set_xlabel('n_estimators')
ax.set_ylabel('learing')
ax.set_zlabel('deep')
plt.title('3D Scatter Plot with Color Mapping')
ax.view_init(elev=40, azim=90)  # 调整这两个参数！
ax.set_zlim(-0.5, 2.5)  # 关键：控制Z轴范围！
plt.show()
'''

import xgboost
print(xgboost.__version__)

##★★★★★★★★★★★★★★★★***************自定义损失函数模型
import numpy as np
from sklearn.base import RegressorMixin
from sklearn.utils.validation import check_X_y, check_array
import torch
import xgboost as xgb
from xgboost.callback import TrainingCallback
from sklearn.utils import check_random_state

class ResidualLossCallback(TrainingCallback):
    """自定义回调类，用于记录每轮的残差损失"""
    def __init__(self, X, y, z, model):
        self.X = X
        self.y = y
        self.z = z
        self.model = model  # 主模型引用

    def after_iteration(self, model, epoch, evals_log):
        """每轮迭代后调用"""
        preds = model.predict(xgb.DMatrix(self.X))
        loss = self.model._compute_residual_loss(self.y, preds, self.z)
        self.model.tree_losses_.append(loss.item())
        return False  # 不提前停止训练

class ResidualLossXGBoost(RegressorMixin):
    """使用XGBoost实现的自定义残差损失回归模型"""
    
    def __init__(self, alpha=0.7, n_estimators=100, max_depth=None, 
                 learning_rate=0.1, subsample=1.0, colsample_bytree=1.0,
                 random_state=None):
        self.alpha = alpha
        self.n_estimators = n_estimators
        self.max_depth = max_depth
        self.learning_rate = learning_rate
        self.subsample = subsample
        self.colsample_bytree = colsample_bytree
        self.random_state = random_state
        self.model_ = None
        self.tree_losses_ = []  # 存储每轮的损失
    
    def _compute_residual_loss(self, y_true, y_pred, z_true):
        """计算残差损失（PyTorch实现）"""
        y_true_t = torch.from_numpy(y_true)
        z_true_t = torch.from_numpy(z_true)
        y_pred_t = torch.from_numpy(y_pred)
        
        loss_y = torch.mean((y_pred_t - y_true_t) ** 2)
        loss_yz = torch.mean((y_pred_t - (y_true_t + z_true_t)) ** 2)
        return self.alpha * loss_y + (1 - self.alpha) * loss_yz
    
    def fit(self, X, y, z):
        """训练模型"""
        X, y = check_X_y(X, y)
        z = check_array(z, ensure_2d=False)
        
        if len(y) != len(z):
            raise ValueError("y和z的长度必须相同")
        
        self.n_features_ = X.shape[1]
        random_state = check_random_state(self.random_state)
        self.tree_losses_ = []
        
        # 准备数据（直接通过闭包传递z）
        dtrain = xgb.DMatrix(X, label=y)
        
        # 自定义目标函数（通过闭包访问z）
        def custom_obj(preds, dtrain):
            y = dtrain.get_label()
            z_val = z  # 直接从外部作用域获取z（闭包）
            
            # 计算梯度
            preds_t = torch.from_numpy(preds)
            y_t = torch.from_numpy(y)
            z_t = torch.from_numpy(z_val)
            
            grad_y = 2 * (preds_t - y_t)
            grad_yz = 2 * (preds_t - (y_t + z_t))
            grad = self.alpha * grad_y + (1 - self.alpha) * grad_yz
            
            # 计算Hessian
            hess = 2 * torch.ones_like(preds_t)
            
            return grad.numpy(), hess.numpy()
        
        # 训练参数
        params = {
            'objective': 'reg:squarederror',  # 基础目标（实际使用custom_obj覆盖）
            'max_depth': self.max_depth,
            'learning_rate': self.learning_rate,
            'subsample': self.subsample,
            'colsample_bytree': self.colsample_bytree,
            'seed': self.random_state,
            'verbosity': 0
        }
        
        # 训练模型（使用自定义回调和目标函数）
        self.model_ = xgb.train(
            params,
            dtrain,
            num_boost_round=self.n_estimators,
            obj=custom_obj,
            callbacks=[ResidualLossCallback(X, y, z, self)]  # 使用TrainingCallback
        )
        return self
    
    def predict(self, X):
        """预测主目标y"""
        X = check_array(X)
        if self.model_ is None:
            raise RuntimeError("模型尚未训练")
        return self.model_.predict(xgb.DMatrix(X))


model = ResidualLossXGBoost(
        alpha=-1,
        #objective='reg:squarederror',  # 回归任务的目标函数
        n_estimators=80,              # 树的数量
        learning_rate=0.1,             # 学习率
        max_depth=2, 
    )

model.fit(X_train.astype(np.float64), y_train.astype(np.float64), z_train.astype(np.float64))
re_test_predictions = model.predict(X_test)
plt.scatter(y_test, re_test_predictions)
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')

mean_absolute_error(y_test, re_test_predictions)################################MAE--5.161
mean_squared_error(y_test, re_test_predictions)#train_mse#######################MSE--40.88
corr2, p_value2 = spearmanr(y_test, re_test_predictions)########################相关0.834
corr2
r2_score(y_test, re_test_predictions)###########################################R2=0.568
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
test_result = re_test_predictions-y_test
    # 找到小于50的样本的索引
indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)#############小于50--0.51
    # 找到大于等于50的样本的索引
indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)##############大于50---0.52



##测试不同的超参
alpha_bc =np.linspace(2, -2, num=41)  # 1 → 0，共11项
bianhua=np.ones((12, len(alpha_bc)))

for i in range(len(alpha_bc)):  # range(20) 生成 0~19
    print(i)
    #model3 = ResidualLossRandomForest(alpha=alpha_bc[i],n_estimators=120, max_depth=7)
    model3 = ResidualLossXGBoost(
        alpha=alpha_bc[i],
        #objective='reg:squarederror',  # 回归任务的目标函数
        n_estimators=80,              # 树的数量
        learning_rate=0.1,             # 学习率
        max_depth=2, 
    )
    model3.fit(X_train.astype(np.float64), y_train.astype(np.float64), z_train.astype(np.float64))
    re_test_predictions = model3.predict(X_test)
    
    bianhua[0,i]=mean_absolute_error(y_test, re_test_predictions)
    bianhua[1,i]=mean_squared_error(y_test, re_test_predictions)#train_mse
    bianhua[2,i]=mean_squared_error(y_test, re_test_predictions)**0.5
    #corr2, p_value2 = spearmanr(y_test, re_test_predictions)
    #bianhua[2,i]=corr2
    bianhua[3,i]=np.corrcoef(np.array(y_test, dtype=float), re_test_predictions.ravel() )[0, 1]
    bianhua[4,i]=r2_score(y_test, re_test_predictions)
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
    test_result = re_test_predictions-y_test
    # 找到小于50的样本的索引
    indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
    bianhua[5,i]=np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)
    # 找到大于等于50的样本的索引
    indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
    bianhua[6,i]=np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)
 


    re_train_predictions_male = model3.predict(X_male)
    #plt.scatter(Y_male,re_train_predictions_male,s=10)
    fenxi_=re_train_predictions_male-Y_male
    from sklearn.metrics import confusion_matrix
    #y_true = [1, 0, 1, 1, 0, 1, 0, 0]
    #y_pred = [1, 0, 0, 1, 1, 1, 0, 0]
    #cm = confusion_matrix(y_true, y_pred)
    #Target_male.shape
    #(fenxi_ > 0).astype(int).reshape(-1).shape
    cm = confusion_matrix(Target_male.astype(int), (fenxi_ > 0).astype(int).reshape(-1))
    # 提取混淆矩阵元素
    TN, FP = cm[0,0], cm[0,1]
    FN, TP = cm[1,0], cm[1,1]
    # 计算指标
    accuracy = (TP + TN) / (TP + TN + FP + FN)
    precision = TP / (TP + FP)
    recall = TP / (TP + FN)
    f1 = 2 * (precision * recall) / (precision + recall)

    bianhua[7,i]=accuracy
    bianhua[8,i]=precision
    bianhua[9,i]=recall
    bianhua[10,i]=f1
    
    # Sigmoid 转换
    sigmoid_arr = 1 / (1 + np.exp(-fenxi_.astype(np.float64)))
    #plt.plot(sigmoid_arr.flatten(),sigmoid_arr.flatten(), color='red', linewidth=1.5, linestyle='-')
    # 转换数据格式
    probabilities = sigmoid_arr.flatten()
    labels = Target_male.astype(np.int32)
    # 计算 AUC
    from sklearn.metrics import roc_auc_score
    auc = roc_auc_score(labels, probabilities)
    bianhua[11,i]=auc
    
type(bianhua)

from openpyxl import load_workbook
df = pd.DataFrame(bianhua)
# 加载现有文件（如果存在）
# 定义目标目录和文件名
#output_dir = r"C:/Users/32684/.spyder-py3/20250418自定义损失"  # Windows 路径示例（注意开头的 r）
# 1. 定义文件路径
filepath = "C:/Users/32684/.spyder-py3/20250418自定义损失/XGBoost_Female.xlsx"  # 或绝对路径（如 r"C:\path\to\A.xlsx"）

# 2. 将 NumPy 数组转为 DataFrame（确保列数匹配）
df = pd.DataFrame(bianhua)

# 3. 加载现有 Excel 文件
wb = load_workbook(filepath)

# 4. 选择工作表（假设为 Sheet1，按需修改）
ws = wb["Sheet1"]  # 或 wb.active

# 5. 从 B2 开始写入数据（保留其他单元格不变）
for row_idx, row in enumerate(df.values, start=2):  # 从第2行开始
    for col_idx, value in enumerate(row, start=2):   # 从第2列（B列）开始
        ws.cell(row=row_idx, column=col_idx, value=value)

# 6. 保存文件
wb.save(filepath)
print(f"数据已从 B2 开始写入，其他数据未受影响。")














################对陆慕他变化，进行可视化
#MAE--5.161
plt.figure(figsize=(22, 3))  # 调整整体画布大小
plt.subplot(1, 6, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[0,:],s=10)
plt.plot([-2,2], [5.16,5.16], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [4,13], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [4,13], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(4, 13)
plt.xlabel('$\lambda$')
plt.ylabel('MAE')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})
plt.grid(True)

#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#MSE--40.880
plt.subplot(1, 6, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[1,:],s=10)
plt.plot([-2,2], [40.88,40.88], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [30,200], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [30,200], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(20, 200)
plt.xlabel('$\lambda$')
plt.ylabel('MSE')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})
plt.grid(True)

#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.834
plt.subplot(1, 6, 3)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[2,:],s=10)
plt.plot([-2,2], [0.83,0.83], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.4,1.1], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.4,1.1], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.4, 1.2)
plt.xlabel('$\lambda$')
plt.ylabel(r'$\rho$')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})
plt.grid(True)

#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#R2=0.568
plt.subplot(1, 6, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[3,:],s=10)
plt.plot([-2,2], [0.57,0.57], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [-1,0.7], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [-1,0.7], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(-1, 0.8)
plt.xlabel('$\lambda$')
plt.ylabel('$R^2$')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='lower right', bbox_to_anchor=(1, 0),prop={'size': 10})
plt.grid(True)

#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.51
plt.subplot(1, 6, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[4,:],s=10)
plt.plot([-2,2], [0.51,0.51], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.45,0.8], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.45,0.8], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.45, 0.8)
plt.xlabel('$\lambda$')
plt.ylabel('P1')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper left', bbox_to_anchor=(0, 1),prop={'size': 10})
plt.grid(True)

#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.52
plt.subplot(1, 6, 6)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[5,:],s=10)
plt.plot([-2,2], [0.52,0.52], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.0,0.6], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.0,0.6], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.0, 0.6)
plt.xlabel('$\lambda$')
plt.ylabel('P2')
plt.title(f'Female (XGBoost, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='lower left', bbox_to_anchor=(0, 0),prop={'size': 10})
plt.grid(True)
plt.subplots_adjust(wspace=0.4)  # 增加水平间距（默认0.2）
plt.show()


#混淆矩阵★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★★
modelaa =ResidualLossXGBoost(
        alpha=0.5,
        #objective='reg:squarederror',  # 回归任务的目标函数
        n_estimators=80,              # 树的数量
        learning_rate=0.1,             # 学习率
        max_depth=2, 
    )
modelaa.fit(X_train.astype(np.float64), y_train.astype(np.float64), z_train.astype(np.float64))
  

re_train_predictions_male = modelaa.predict(X_male)

#plt.scatter(Y_male,re_train_predictions_male,s=10)

fenxi_=re_train_predictions_male-Y_male
from sklearn.metrics import confusion_matrix
#y_true = [1, 0, 1, 1, 0, 1, 0, 0]
#y_pred = [1, 0, 0, 1, 1, 1, 0, 0]
#cm = confusion_matrix(y_true, y_pred)
#Target_male.shape
#(fenxi_ > 0).astype(int).reshape(-1).shape
cm = confusion_matrix(Target_male.astype(int), (fenxi_ > 0).astype(int).reshape(-1))
# 提取混淆矩阵元素
TN, FP = cm[0,0], cm[0,1]
FN, TP = cm[1,0], cm[1,1]

# 计算指标
accuracy = (TP + TN) / (TP + TN + FP + FN)
precision = TP / (TP + FP)
recall = TP / (TP + FN)
f1 = 2 * (precision * recall) / (precision + recall)









'''
X = np.random.rand(1000, 10)
y = 10 * X[:, 0] + 5 * X[:, 1] + np.random.normal(0, 1, 1000)
z = 2 * X[:, 2] + 0.5 * X[:, 3] + np.random.normal(0, 0.5, 1000)
type(X)
model.fit(X, y.astype(np.float64), z.astype(np.float64))

X = np.random.rand(100, 5)
y = np.random.rand(100)
z = np.random.rand(100)
    
    # 训练模型
model = ResidualLossXGBoost(
        alpha=0.7,
        n_estimators=80,
        learning_rate=0.1,
        max_depth=2,
        random_state=42
    )
model.fit(X, y, z)
'''




