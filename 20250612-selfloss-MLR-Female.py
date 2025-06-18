import numpy as np
from sklearn.model_selection import train_test_split
from sklearn.preprocessing import StandardScaler
import pandas as pd
from sklearn.metrics import mean_squared_error
from sklearn.metrics import mean_absolute_error  # 正确导入方式
from sklearn.metrics import r2_score  # 必须导入！
import matplotlib.pyplot as plt
import seaborn as sns

from scipy.stats import spearmanr
from scipy.stats import pearsonr

# 11111-----------------------生成数据 (1000个样本，5个特征)

DATA = pd.read_excel('C:/Users/32684/.spyder-py3/20250418自定义损失/Atherosclerosis_data.xlsx', header=None)

DATA1 = DATA.values
print(DATA1)

Y=DATA1[1:,1]
print(Y)

#X=DATA1[1:,[3,4]]
X1=DATA1[1:,2:-1]
print(X1)

print(DATA1[0,2:-1])

X=X1[:,[0,1,2,4,7]]
'''
from sklearn.linear_model import LassoCV
from sklearn.preprocessing import StandardScaler
import numpy as np

# 假设 X 是特征矩阵，y 是年龄
scaler = StandardScaler()
X_scaled = scaler.fit_transform(X1)  # 标准化特征

# 使用交叉验证选择最优 lambda
lasso = LassoCV(cv=5).fit(X_scaled, Y)

# 特征重要性：系数绝对值
importance = np.abs(lasso.coef_)
feature_names = [...]  # 你的特征名称列表

##★★★★★★★★★★★★★★★★★★★★★置信区间
from sklearn.utils import resample

def bootstrap_ci(X, y, model, n_bootstraps=1000, alpha=0.05):
    coefs = []
    for _ in range(n_bootstraps):
        X_resampled, y_resampled = resample(X, y)
        model.fit(X_resampled, y_resampled)
        coefs.append(model.coef_)
    
    coefs = np.array(coefs)
    lower = np.percentile(coefs, (alpha/2)*100, axis=0)
    upper = np.percentile(coefs, (1-alpha/2)*100, axis=0)
    return lower, upper

# 计算95%置信区间
lower_ci, upper_ci = bootstrap_ci(X_scaled, Y, lasso)

##★★★★★★★★★★★★★★★★★★★★★p值
from sklearn.base import clone

def permutation_test(X, y, model, n_permutations=1000):
    observed_coef = model.fit(X, y).coef_
    perm_coefs = np.zeros((n_permutations, X.shape[1]))
    
    for i in range(n_permutations):
        y_permuted = np.random.permutation(y)
        perm_model = clone(model)
        perm_coefs[i] = perm_model.fit(X, y_permuted).coef_
    
    p_values = np.mean(np.abs(perm_coefs) >= np.abs(observed_coef), axis=0)
    return p_values

# 计算p值
p_values = permutation_test(X_scaled, Y, lasso)
'''

Sex=DATA1[1:,0]
Target=DATA1[1:,-1]

X_male= X[np.where(Sex == 0)[0],:]
Y_male= Y[np.where(Sex == 0)[0]]
Target_male=Target[np.where(Sex == 0)[0]]

X_male_h = X_male[np.where(Target_male == 0)[0],:]
Y_male_h = Y_male[np.where(Target_male == 0)[0]]
X_male_noh = X_male[np.where(Target_male == 1)[0],:]
Y_male_noh = Y_male[np.where(Target_male == 1)[0]]

'''
X_female= X[np.where(Sex == 0)[0],:]
Y_female= Y[np.where(Sex == 0)[0]]
Target_female=Target[np.where(Sex == 0)[0]]
'''

#★★★★★★★★★★★★★★★★★★★★★★★现在测试男性的数据
X_model=X_male_h
Y_model=Y_male_h
#print(np.size(Y_model))


X_train=X_model[::2]  # 从索引0开始，步长为2
y_train=Y_model[::2]

X_test=X_model[1::2]  # 从索引1开始，步长为2 
y_test=Y_model[1::2]


#----------------第一步预训练构建z,相对个体风险差异
from sklearn.linear_model import LinearRegression
model1 = LinearRegression()
model1.fit(X_train, y_train)

#构造训练集预测结果，和测试集预测结果
train_predictions = model1.predict(X_train)
test_predictions = model1.predict(X_test)

#model2是基于X_train和y_train，构造mpa模型
model2 = LinearRegression()
model2.fit(y_train.reshape(-1, 1),train_predictions)
#构造训练集的mpa
train_predictions_mpa = model2.predict(y_train.reshape(-1, 1))
#构造测试集的mpa
test_predictions_mpa = model2.predict(y_test.reshape(-1, 1))

z_train=train_predictions-train_predictions_mpa
z_test=test_predictions-test_predictions_mpa


'''
import sys
sys.path.append(r"C:/Users/32684/.spyder-py3/20250418自定义损失")  # 添加自定义模块的路径
from ResidualLossRandomForest import ResidualLossRandomForest
'''

'''
#经测试，单独使用pa或者mpa都不能作为，新标签
#from sklearn.ensemble import RandomForestRegressor
model3 = ResidualLossRandomForest()
model3.fit(X_train, train_predictions-train_predictions_mpa+y_train)
#model3.fit(X_train, train_predictions_mpa)
pre = model3.predict(X_train)
plt.scatter(y_train, pre,s=10,color='b')
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')

'''
mean_absolute_error(y_test, test_predictions)################################MAE--5.161
mean_squared_error(y_test, test_predictions)#train_mse#######################MSE--40.88
r=np.corrcoef(np.array(y_test, dtype=float), test_predictions.ravel() )[0, 1]
r2_score(y_test, test_predictions)###########################################R2=0.568
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
test_result = test_predictions-y_test
    # 找到小于50的样本的索引
indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)#############小于50--0.51
    # 找到大于等于50的样本的索引
indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)##############大于50---0.52


plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
plt.subplot(2, 3, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions,s=10,color='b', alpha=0.2)
corr_coef = np.corrcoef(np.array(y_train, dtype=float), train_predictions.ravel() )[0, 1]
#mean_squared_error(y_train, train_predictions)
modela = LinearRegression()
modela.fit(y_train.reshape(-1, 1), train_predictions)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}\nr = {corr_coef:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Arterial biological age')
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions-y_train,s=10,color='b', alpha=0.2)
modela.fit(y_train.reshape(-1, 1), train_predictions-y_train)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
#plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Residual')
plt.xlim(20, 90)
plt.ylim(-35, 35)
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 3)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(train_predictions-y_train, kde=True, bins=20, color='b', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Pre-Residual')
plt.xlim(-35, 35)
plt.ylim(0, 0.075)
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions,s=10,color='c', alpha=0.2)
corr_coef = np.corrcoef(np.array(y_train, dtype=float), train_predictions.ravel() )[0, 1]
modela.fit(y_test.reshape(-1, 1), test_predictions)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}\nr = {corr_coef:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Pre-Arterial biological age')
plt.title(f'Female (MLR, test, n={len(y_test)})')

plt.subplot(2, 3, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions-y_test,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions-y_test)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
#plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlim(20, 90)
plt.xlabel('Chronological age')
plt.ylabel('Pre-Residual')
plt.xlim(20, 90)
plt.ylim(-35, 35)
plt.title(f'Female (MLR, test, n={len(y_test)})')

plt.subplot(2, 3, 6)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(test_predictions-y_test, kde=True, bins=20, color='c', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Pre-Residual')
plt.xlim(-35, 35)
plt.ylim(0, 0.075)
plt.title(f'Female (MLR, test, n={len(y_test)})')
plt.tight_layout()
plt.show()


#############     混淆矩阵分析
model1 = LinearRegression()
model1.fit(X_train, y_train)
male_predictions = model1.predict(X_male)
fenxi_=male_predictions-Y_male
from sklearn.metrics import confusion_matrix
#y_true = [1, 0, 1, 1, 0, 1, 0, 0]
#y_pred = [1, 0, 0, 1, 1, 1, 0, 0]
#cm = confusion_matrix(y_true, y_pred)
#Target_male.shape
#(fenxi_ > 0).astype(int).reshape(-1).shape
cm = confusion_matrix(Target_male.astype(int), (fenxi_ > 0).astype(int).reshape(-1))
# 提取混淆矩阵元素
TN, FP = cm[0,0], cm[0,1]
FN, TP = cm[1,0], cm[1,1]

# 计算指标
accuracy = (TP + TN) / (TP + TN + FP + FN)
precision = TP / (TP + FP)
recall = TP / (TP + FN)
f1 = 2 * (precision * recall) / (precision + recall)


#######--------------L1
model11 = LinearRegression()
model11.fit(X_train, y_train+z_train)
#构造训练集预测结果，和测试集预测结果
train_predictions_l1 = model11.predict(X_train)
test_predictions_l1 = model11.predict(X_test)


mean_absolute_error(y_test, test_predictions_l1)################################MAE--5.161
mean_squared_error(y_test, test_predictions_l1)#train_mse#######################MSE--40.88
corr2, p_value2 = spearmanr(y_test, test_predictions_l1)########################相关0.834
corr2
r2_score(y_test, test_predictions_l1)###########################################R2=0.568
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
test_result = test_predictions_l1-y_test
    # 找到小于50的样本的索引
indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)#############小于50--0.51
    # 找到大于等于50的样本的索引
indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)##############大于50---0.52



plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
plt.subplot(2, 3, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions_l1,s=10,color='b', alpha=0.2)
#mean_squared_error(y_train, train_predictions)
modela = LinearRegression()
modela.fit(y_train.reshape(-1, 1), train_predictions_l1)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Re-Arterial age')
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(y_train, train_predictions_l1-y_train,s=10,color='b', alpha=0.2)
modela.fit(y_train.reshape(-1, 1), train_predictions_l1-y_train)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlabel('Chronological age')
plt.ylabel('Re-Residual')
plt.xlim(20, 90)
plt.ylim(-30, 30)
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 3)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(train_predictions_l1-y_train, kde=True, bins=20, color='b', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Re-Residual')
plt.xlim(-30, 30)
plt.ylim(0, 0.075)
plt.title(f'Female (MLR, train, n={len(y_train)})')

plt.subplot(2, 3, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions_l1,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions_l1)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
a = modela.coef_[0]  # 斜率
b = modela.intercept_  # 截距
equation = f'y = {a:.2f}x + {b:.2f}'
plt.text(0.05, 0.95, equation, transform=plt.gca().transAxes, 
         fontsize=13, color='red', verticalalignment='top', bbox=dict(facecolor='white', alpha=0.8))
plt.xlim(20, 90)
plt.ylim(20, 90)
plt.plot([20,90], [20,90], color='black', linewidth=1, linestyle='--', label='y=x')
#plt.plot([20,90], [20,90], color='red', linewidth=1, linestyle='-', label='y=x')
plt.xlabel('Chronological age')
plt.ylabel('Re-Arterial age')
plt.title(f'Female (MLR, test, n={len(y_test)})')

plt.subplot(2, 3, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(y_test, test_predictions_l1-y_test,s=10,color='c', alpha=0.2)
modela.fit(y_test.reshape(-1, 1), test_predictions_l1-y_test)
predia = modela.predict(np.array([25,85]).reshape(-1, 1))
plt.plot([25,85], predia, color='red', linewidth=1.5, linestyle='-')
plt.xlim(20, 90)
plt.xlabel('Chronological age')
plt.ylabel('Re-Residual')
plt.xlim(20, 90)
plt.ylim(-30, 30)
plt.title(f'Female (MLR, test, n={len(y_test)})')

plt.subplot(2, 3, 6)  # (行数, 列数, 当前子图索引)
sns.set(style='whitegrid')
sns.distplot(test_predictions_l1-y_test, kde=True, bins=20, color='c', hist_kws={'edgecolor':'black'})
plt.ylabel('Probability')
plt.xlabel('Re-Residual')
plt.xlim(-30, 30)
plt.ylim(0, 0.075)
plt.title(f'Female (MLR, test, n={len(y_test)})')
plt.tight_layout()
plt.show()

###混淆矩阵测试
model11 = LinearRegression()
model11.fit(X_train, y_train+z_train)
male_predictions = model11.predict(X_male)
fenxi_=male_predictions-Y_male
from sklearn.metrics import confusion_matrix
#y_true = [1, 0, 1, 1, 0, 1, 0, 0]
#y_pred = [1, 0, 0, 1, 1, 1, 0, 0]
#cm = confusion_matrix(y_true, y_pred)
#Target_male.shape
#(fenxi_ > 0).astype(int).reshape(-1).shape
cm = confusion_matrix(Target_male.astype(int), (fenxi_ > 0).astype(int).reshape(-1))
# 提取混淆矩阵元素
TN, FP = cm[0,0], cm[0,1]
FN, TP = cm[1,0], cm[1,1]

# 计算指标
accuracy = (TP + TN) / (TP + TN + FP + FN)
precision = TP / (TP + FP)
recall = TP / (TP + FN)
f1 = 2 * (precision * recall) / (precision + recall)

#---------------------------------预训练结束，获得Z 

# 标准化数据（推荐）
scaler = StandardScaler()
X_train = scaler.fit_transform(X_train)
X_test = scaler.transform(X_test)
X_male = scaler.transform(X_male)

# 转换为 PyTorch 张量
import torch

X_train_tensor = torch.FloatTensor(X_train)
#y_train_tensor = torch.FloatTensor(y_train)
y_train_tensor = torch.FloatTensor(y_train.astype(np.float32))
z_train_tensor = torch.FloatTensor(z_train.astype(np.float32))

X_male_tensor = torch.FloatTensor(X_male)

X_test_tensor = torch.FloatTensor(X_test)
#y_test_tensor = torch.FloatTensor(y_test)
y_test_tensor = torch.FloatTensor(y_test.astype(np.float32))
z_test_tensor = torch.FloatTensor(z_test.astype(np.float32))

# 22222222222-----------------------定义模型 
import torch.nn as nn

class LinearRegression(nn.Module):
    def __init__(self, input_dim):
        super(LinearRegression, self).__init__()
        self.linear = nn.Linear(input_dim, 1)  # 输入维度 → 输出1个值
    
    def forward(self, x):
        return self.linear(x)

# 初始化模型
model = LinearRegression(input_dim=X_train.shape[1])  # 5个特征

# 333333333333-----------------------定义损失函数和优化器 
  
alpha_bc =np.linspace(2, -2, num=41)  # 1 → 0，共11项
bianhua=np.ones((12, len(alpha_bc)))

for i in range(len(alpha_bc)):  # range(20) 生成 0~19
    print(i)
    #i=20
    
    def residual_loss(y_pred, y_true, z_true, alpha=alpha_bc[i]):
        loss_y = torch.mean((y_pred - y_true) ** 2)  # 主目标损失
        loss_yz = torch.mean((y_pred - (y_true + z_true)) ** 2)  # 次目标损失（与主目标共享预测值）
        return alpha * loss_y + (1 - alpha) * loss_yz
    '''
    def residual_loss(y_pred, y_true, z_true, alpha=alpha_bc[i]):
        # Log-Cosh Loss 替代 MSE
        def log_cosh(x):
            return torch.log(torch.cosh(x))
        # 主目标损失（Log-Cosh）
        loss_y = torch.mean(log_cosh(y_pred - y_true))
        # 次目标损失（Log-Cosh）
        loss_yz = torch.mean(log_cosh(y_pred - (y_true + z_true)))
        # 加权总损失
        return alpha * loss_y + (1 - alpha) * loss_yz
    '''
    # 修改点3：训练循环调整
    optimizer = torch.optim.SGD(model.parameters(), lr=0.01)
    num_epochs = 200
    train_losses = []
    test_losses = []

    for epoch in range(num_epochs):
        model.train()
        outputs = model(X_train_tensor)
    
        # 修改点4：使用residual_loss计算损失
        loss = residual_loss(outputs.squeeze(), y_train_tensor, z_train_tensor)
    
        optimizer.zero_grad()
        loss.backward()
        optimizer.step()
        train_losses.append(loss.item())
    
        model.eval()
        with torch.no_grad():
            test_outputs = model(X_test_tensor)
            test_loss = residual_loss(test_outputs.squeeze(), y_test_tensor, z_test_tensor)
            test_losses.append(test_loss.item())
    
        if (epoch + 1) % 10 == 0:
            print(f'Epoch [{epoch+1}/{num_epochs}], Loss: {loss.item():.4f}, Test Loss: {test_loss.item():.4f}')

    # 修改点5：评估时计算两种指标
    model.eval()
    with torch.no_grad():
        re_test_predictions = model(X_test_tensor).squeeze().numpy()
    
    with torch.no_grad():
        re_train_predictions = model(X_train_tensor).squeeze().numpy()
         
    with torch.no_grad():
        re_train_predictions_male = model(X_male_tensor).squeeze().numpy()
    
    bianhua[0,i]=mean_absolute_error(y_test, re_test_predictions)
    bianhua[1,i]=mean_squared_error(y_test, re_test_predictions)#train_mse
    #corr2, p_value2 = spearmanr(y_test, re_test_predictions)
    bianhua[2,i]=mean_squared_error(y_test, re_test_predictions)**0.5
    bianhua[3,i]=np.corrcoef(np.array(y_test, dtype=float), re_test_predictions.ravel() )[0, 1]
    bianhua[4,i]=r2_score(y_test, re_test_predictions)
    #corr1, p_value1 = spearmanr(y_train, re_train_predictions-y_train)
    #bianhua[4,i]=corr1
    test_result = re_test_predictions-y_test
    # 找到小于50的样本的索引
    indices_mi = np.where(y_test.astype(int) < 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
    bianhua[5,i]=np.sum(test_result[indices_mi].astype(float) > 0)/len(indices_mi)
    # 找到大于等于50的样本的索引
    indices_ma = np.where(y_test.astype(int) >= 50)[0]
    #len(indices_mi)
    # 计算大于0的样本数量
    #np.sum(train_result[indices_mi].astype(float) > 0)
    bianhua[6,i]=np.sum(test_result[indices_ma].astype(float) > 0)/len(indices_ma)
    
    fenxi_=re_train_predictions_male-Y_male
    from sklearn.metrics import confusion_matrix
    #y_true = [1, 0, 1, 1, 0, 1, 0, 0]
    #y_pred = [1, 0, 0, 1, 1, 1, 0, 0]
    #cm = confusion_matrix(y_true, y_pred)
    #Target_male.shape
    #(fenxi_ > 0).astype(int).reshape(-1).shape
    cm = confusion_matrix(Target_male.astype(int), (fenxi_ > 0).astype(int).reshape(-1))
    # 提取混淆矩阵元素
    TN, FP = cm[0,0], cm[0,1]
    FN, TP = cm[1,0], cm[1,1]

    # 计算指标
    accuracy = (TP + TN) / (TP + TN + FP + FN)
    precision = TP / (TP + FP)
    recall = TP / (TP + FN)
    f1 = 2 * (precision * recall) / (precision + recall)
    bianhua[7,i]=accuracy
    bianhua[8,i]=precision
    bianhua[9,i]=recall
    bianhua[10,i]=f1
    
    # Sigmoid 转换
    sigmoid_arr = 1 / (1 + np.exp(-fenxi_.astype(np.float64)))
    #plt.plot(sigmoid_arr.flatten(),sigmoid_arr.flatten(), color='red', linewidth=1.5, linestyle='-')
    # 转换数据格式
    probabilities = sigmoid_arr.flatten()
    labels = Target_male.astype(np.int32)
    # 计算 AUC
    from sklearn.metrics import roc_auc_score
    auc = roc_auc_score(labels, probabilities)
    bianhua[11,i]=auc
    
type(bianhua)

from openpyxl import load_workbook
df = pd.DataFrame(bianhua)
# 加载现有文件（如果存在）
# 定义目标目录和文件名
#output_dir = r"C:/Users/32684/.spyder-py3/20250418自定义损失"  # Windows 路径示例（注意开头的 r）
# 1. 定义文件路径
filepath = "C:/Users/32684/.spyder-py3/20250418自定义损失/MLR_Female.xlsx"  # 或绝对路径（如 r"C:\path\to\A.xlsx"）

# 2. 将 NumPy 数组转为 DataFrame（确保列数匹配）
df = pd.DataFrame(bianhua)

# 3. 加载现有 Excel 文件
wb = load_workbook(filepath)

# 4. 选择工作表（假设为 Sheet1，按需修改）
ws = wb["Sheet1"]  # 或 wb.active

# 5. 从 B2 开始写入数据（保留其他单元格不变）
for row_idx, row in enumerate(df.values, start=2):  # 从第2行开始
    for col_idx, value in enumerate(row, start=2):   # 从第2列（B列）开始
        ws.cell(row=row_idx, column=col_idx, value=value)

# 6. 保存文件
wb.save(filepath)
print(f"数据已从 B2 开始写入，其他数据未受影响。")

    
################对陆慕他变化，进行可视化
#MAE--5.161
plt.figure(figsize=(22, 3))  # 调整整体画布大小
plt.subplot(2, 5, 1)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[0,:],s=10)
plt.plot([-2,2], [5.16,5.16], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [4,13], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [4,13], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(4, 13)
plt.xlabel('$\lambda$')
plt.ylabel('MAE')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})


#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#MSE--40.880
plt.subplot(2, 5, 2)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[1,:],s=10)
plt.plot([-2,2], [40.88,40.88], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [30,200], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [30,200], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(20, 200)
plt.xlabel('$\lambda$')
plt.ylabel('MSE')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})


#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.834
plt.subplot(2, 5, 3)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[2,:],s=10)
plt.plot([-2,2], [0.83,0.83], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.4,1.1], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.4,1.1], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.4, 1.2)
plt.xlabel('$\lambda$')
plt.ylabel(r'$\rho$')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper right', bbox_to_anchor=(1, 1),prop={'size': 10})


#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#R2=0.568
plt.subplot(2, 5, 4)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[3,:],s=10)
plt.plot([-2,2], [0.57,0.57], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [-1,0.7], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [-1,0.7], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(-1, 0.8)
plt.xlabel('$\lambda$')
plt.ylabel('$R^2$')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='lower right', bbox_to_anchor=(1, 0),prop={'size': 10})



#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.51
plt.subplot(2, 5, 5)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[4,:],s=10)
plt.plot([-2,2], [0.51,0.51], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.45,0.8], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.45,0.8], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.45, 0.8)
plt.xlabel('$\lambda$')
plt.ylabel('P1')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='upper left', bbox_to_anchor=(0, 1),prop={'size': 10})


#plt.figure(figsize=(8.8, 6))  # 调整整体画布大小
#0.52
plt.subplot(2, 5, 6)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[5,:],s=10)
plt.plot([-2,2], [0.52,0.52], color='black', linewidth=1.5, linestyle='--', label='KDM')
plt.plot([1,1], [0.0,0.6], color='red', linewidth=1.5, linestyle='-.', label='Loss(MSE)')
plt.plot([0,0], [0.0,0.6], color='red', linewidth=1.5, linestyle='-', label='Loss1')
plt.xlim(-2.1, 2.1)
plt.ylim(0.0, 0.6)
plt.xlabel('$\lambda$')
plt.ylabel('P2')
plt.title(f'Female (MLR, test, n={len(y_train)})')
plt.xticks([-2, -1.5, -1, -0.5, 0, 0.5, 1, 1.5, 2], rotation=45)  # 指定要显示的刻度值
plt.legend(loc='lower left', bbox_to_anchor=(0, 0),prop={'size': 10})
plt.subplots_adjust(wspace=0.4)  # 增加水平间距（默认0.2）


plt.subplot(2, 5, 7)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[6,:],s=10)

plt.subplot(2, 5, 8)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[7,:],s=10)

plt.subplot(2, 5, 9)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[8,:],s=10)

plt.subplot(2, 5, 10)  # (行数, 列数, 当前子图索引)
plt.scatter(alpha_bc,bianhua[9,:],s=10)
plt.show()
